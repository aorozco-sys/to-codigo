"""Unit tests for the audit tracking feature (report-as-source-of-truth).

Covers:
- ``read_previous_report`` for XLSX, CSV, JSON, and missing files.
- ``compare_files`` — unchanged, modified (size/mtime), new, removed.
- ``calculate_audit_stats`` — correct LOC counts.
- ``apply_audit_to_files`` — FileInfo gets correct audit_status and audit_marked.
- Full roundtrip: generate XLSX → manually edit → regenerate → verify marks.
- Reporter audit columns: CSV, XLSX, JSON, HTML.
- Backward compatibility: legacy ``AuditState`` still works.
- ``--audit`` NOT used: no audit columns in output.
"""

from __future__ import annotations

import json
import os
import time
from dataclasses import replace

from to_codigo.core.models import (
    AuditEntry,
    AuditState,
    FileInfo,
    FileChangeStatus,
    AuditDiff,
)
from to_codigo.core.audit import (
    read_previous_report,
    compare_files,
    calculate_audit_stats,
    apply_audit_to_files,
    load_audit_state,
    save_audit_state,
    mark_file,
    merge_audit_state,
)
from to_codigo.core.reporter import CSVReporter, XLSXReporter, JSONReporter, HTMLReporter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_fileinfo(
    path: str,
    loc: int = 100,
    size: int = 0,
    mtime: str = "2025-01-01 00:00:00",
) -> FileInfo:
    """Create a minimal ``FileInfo`` for testing."""
    return FileInfo(
        absolute_path=path,
        relative_path=".",
        filename=path.rsplit("/", 1)[-1],
        language="Python",
        size_bytes=size,
        modified_at=mtime,
        total_lines=loc,
        code_lines=loc,
        comment_lines=0,
        blank_lines=0,
    )


# ---------------------------------------------------------------------------
# read_previous_report
# ---------------------------------------------------------------------------

def test_read_previous_report_missing_file(tmp_path):
    """``read_previous_report`` returns empty dict for missing file."""
    result = read_previous_report(str(tmp_path / "nonexistent.xlsx"))
    assert result == {}


def test_read_previous_report_xlsx(tmp_path):
    """``read_previous_report`` reads XLSX with audit columns."""
    from openpyxl import Workbook

    path = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws = wb.active
    # Write headers matching HEADERS + AUDIT_HEADERS
    headers = [
        "Ruta_Absoluta_Archivo", "Ruta_Relativa_Carpeta", "Nombre_Archivo",
        "Lenguaje_Programacion", "Tamano_Bytes", "Fecha_Modificacion",
        "Lineas_Totales", "Lineas_Codigo", "Lineas_Comentarios", "Lineas_Blanco",
        "TODOs", "FIXMEs", "Auditado", "Estado",
    ]
    ws.append(headers)
    ws.append(["/tmp/a.py", ".", "a.py", "Python", 100, "2025-01-01 00:00:00", 10, 10, 0, 0, 0, 0, "Si", "Auditado"])
    ws.append(["/tmp/b.py", ".", "b.py", "Python", 200, "2025-01-01 00:00:00", 20, 20, 0, 0, 0, 0, "No", "Pendiente"])
    wb.save(path)

    result = read_previous_report(path)
    assert "/tmp/a.py" in result
    assert result["/tmp/a.py"]["audited"] == "Si"
    assert result["/tmp/a.py"]["size"] == 100
    assert result["/tmp/a.py"]["mtime"] == "2025-01-01 00:00:00"
    assert result["/tmp/b.py"]["audited"] == "No"


def test_read_previous_report_csv(tmp_path):
    """``read_previous_report`` reads CSV with audit columns."""
    import csv

    path = str(tmp_path / "test.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Ruta_Absoluta_Archivo", "Tamano_Bytes", "Fecha_Modificacion", "Auditado",
        ])
        writer.writerow(["/tmp/a.py", 100, "2025-01-01 00:00:00", "Si"])
        writer.writerow(["/tmp/b.py", 200, "2025-01-01 00:00:00", "No"])

    result = read_previous_report(path)
    assert "/tmp/a.py" in result
    assert result["/tmp/a.py"]["audited"] == "Si"
    assert result["/tmp/a.py"]["size"] == 100
    assert result["/tmp/b.py"]["audited"] == "No"


def test_read_previous_report_json(tmp_path):
    """``read_previous_report`` reads JSON report."""
    path = str(tmp_path / "test.json")
    payload = {
        "files": [
            {
                "absolute_path": "/tmp/a.py",
                "size_bytes": 100,
                "modified_at": "2025-01-01 00:00:00",
                "audit_marked": "Si",
                "audit_status": "Auditado",
            },
            {
                "absolute_path": "/tmp/b.py",
                "size_bytes": 200,
                "modified_at": "2025-01-01 00:00:00",
                "audit_marked": "No",
                "audit_status": "Pendiente",
            },
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f)

    result = read_previous_report(path)
    assert "/tmp/a.py" in result
    assert result["/tmp/a.py"]["audited"] == "Si"
    assert result["/tmp/a.py"]["size"] == 100
    assert result["/tmp/b.py"]["audited"] == "No"


# ---------------------------------------------------------------------------
# compare_files
# ---------------------------------------------------------------------------

def test_compare_unchanged_file_keeps_si():
    """File with same size + mtime keeps 'Si'."""
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00")]
    diff = compare_files(files, prev)
    assert diff.audited_unchanged == 1
    assert diff.changes["/tmp/a.py"].status == "Auditado"
    assert diff.changes["/tmp/a.py"].audited == "Si"


def test_compare_different_size_becomes_modificado():
    """File with different size becomes 'Modificado'."""
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=200, mtime="2025-01-01 00:00:00")]
    diff = compare_files(files, prev)
    assert diff.audited_modified == 1
    assert diff.changes["/tmp/a.py"].status == "Modificado"


def test_compare_different_mtime_becomes_modificado():
    """File with different mtime becomes 'Modificado'."""
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-02-01 00:00:00")]
    diff = compare_files(files, prev)
    assert diff.audited_modified == 1
    assert diff.changes["/tmp/a.py"].status == "Modificado"


def test_compare_new_file():
    """File not in previous becomes 'Nuevo'."""
    files = [_make_fileinfo("/tmp/new.py", loc=50)]
    diff = compare_files(files, {})
    assert diff.new_files == 1
    assert diff.changes["/tmp/new.py"].status == "Nuevo"


def test_compare_removed_files():
    """Files in previous but not in current are counted as removed."""
    prev = {
        "/tmp/old.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    files = [_make_fileinfo("/tmp/a.py", loc=100)]
    diff = compare_files(files, prev)
    assert diff.removed_files == 1


def test_compare_pending_unchanged():
    """File marked 'No' and unchanged stays 'Pendiente'."""
    prev = {
        "/tmp/a.py": {"audited": "No", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00")]
    diff = compare_files(files, prev)
    assert diff.pending_unchanged == 1
    assert diff.changes["/tmp/a.py"].status == "Pendiente"


# ---------------------------------------------------------------------------
# calculate_audit_stats
# ---------------------------------------------------------------------------

def test_calculate_audit_stats_correct():
    """``calculate_audit_stats`` returns correct LOC counts."""
    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/b.py", loc=200, size=200, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/c.py", loc=50, size=50, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/d.py", loc=75, size=75, mtime="2025-01-01 00:00:00"),
    ]
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
        "/tmp/b.py": {"audited": "Si", "size": 999, "mtime": "2025-01-01 00:00:00"},  # modified
        "/tmp/c.py": {"audited": "No", "size": 50, "mtime": "2025-01-01 00:00:00"},   # pending
        # /tmp/d.py is new
    }
    diff = compare_files(files, prev)
    stats = calculate_audit_stats(files, diff)

    assert stats["audited_files"] == 1
    assert stats["audited_loc"] == 100
    assert stats["modified_files"] == 1
    assert stats["modified_loc"] == 200
    assert stats["new_files"] == 1
    assert stats["new_loc"] == 75
    assert stats["pending_files"] == 1
    assert stats["pending_loc"] == 50
    assert stats["total_files"] == 4
    assert stats["total_loc"] == 425
    assert abs(stats["pct_audited"] - (100 / 425 * 100)) < 0.01


# ---------------------------------------------------------------------------
# apply_audit_to_files
# ---------------------------------------------------------------------------

def test_apply_audit_sets_status_and_marked():
    """``apply_audit_to_files`` updates FileInfo correctly."""
    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/b.py", loc=200, size=999, mtime="2025-01-01 00:00:00"),
    ]
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
        "/tmp/b.py": {"audited": "Si", "size": 999, "mtime": "2025-01-01 00:00:00"},  # unchanged here
    }
    diff = compare_files(files, prev)
    updated = apply_audit_to_files(files, diff)
    assert updated[0].audit_status == "Auditado"
    assert updated[0].audit_marked == "Si"
    assert updated[1].audit_status == "Auditado"


def test_apply_audit_modified_resets_to_no():
    """Modified files get audit_marked reset to 'No'."""
    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=200, mtime="2025-01-01 00:00:00"),
    ]
    prev = {
        "/tmp/a.py": {"audited": "Si", "size": 100, "mtime": "2025-01-01 00:00:00"},
    }
    diff = compare_files(files, prev)
    updated = apply_audit_to_files(files, diff)
    assert updated[0].audit_status == "Modificado"
    assert updated[0].audit_marked == "No"


# ---------------------------------------------------------------------------
# Full roundtrip: XLSX generate → edit → regenerate
# ---------------------------------------------------------------------------

def test_xlsx_roundtrip_preserves_marks(tmp_path):
    """Generate XLSX → read back → compare → regenerate → marks preserved for unchanged."""
    from to_codigo.core.scanner import compute_stats

    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/b.py", loc=200, size=200, mtime="2025-01-01 00:00:00"),
    ]
    stats = compute_stats(files)
    output = str(tmp_path / "reporte.xlsx")

    # --- First run: all "Nuevo" ---
    diff1 = compare_files(files, {})
    updated1 = apply_audit_to_files(files, diff1)
    audit_stats1 = calculate_audit_stats(updated1, diff1)
    XLSXReporter().generate(updated1, stats, output, audit_state=True, audit_stats=audit_stats1)

    # --- Read back and simulate auditor marking a.py as "Si" ---
    prev_data = read_previous_report(output)
    prev_data["/tmp/a.py"]["audited"] = "Si"

    # Re-write the XLSX with the auditor's marks
    diff1b = compare_files(files, prev_data)
    updated1b = apply_audit_to_files(files, diff1b)
    audit_stats1b = calculate_audit_stats(updated1b, diff1b)
    XLSXReporter().generate(updated1b, stats, output, audit_state=True, audit_stats=audit_stats1b)

    # --- Second run: read previous, compare ---
    prev2 = read_previous_report(output)
    diff2 = compare_files(files, prev2)
    updated2 = apply_audit_to_files(files, diff2)

    # a.py should be "Auditado" with "Si" (unchanged)
    a_info = next(f for f in updated2 if f.absolute_path == "/tmp/a.py")
    assert a_info.audit_status == "Auditado"
    assert a_info.audit_marked == "Si"

    # b.py should be "Pendiente" with "No" (was never marked)
    b_info = next(f for f in updated2 if f.absolute_path == "/tmp/b.py")
    assert b_info.audit_status == "Pendiente"
    assert b_info.audit_marked == "No"


def test_xlsx_roundtrip_modified_resets(tmp_path):
    """Generate XLSX → mark file → change file → regenerate → file is 'Modificado'."""
    from to_codigo.core.scanner import compute_stats

    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00"),
    ]
    stats = compute_stats(files)
    output = str(tmp_path / "reporte2.xlsx")

    # --- First run ---
    diff1 = compare_files(files, {})
    updated1 = apply_audit_to_files(files, diff1)
    audit_stats1 = calculate_audit_stats(updated1, diff1)
    XLSXReporter().generate(updated1, stats, output, audit_state=True, audit_stats=audit_stats1)

    # --- Read back, simulate auditor marking as "Si" ---
    prev_data = read_previous_report(output)
    prev_data["/tmp/a.py"]["audited"] = "Si"
    diff1b = compare_files(files, prev_data)
    updated1b = apply_audit_to_files(files, diff1b)
    audit_stats1b = calculate_audit_stats(updated1b, diff1b)
    XLSXReporter().generate(updated1b, stats, output, audit_state=True, audit_stats=audit_stats1b)

    # --- Second run: file has CHANGED (different size) ---
    changed_files = [
        _make_fileinfo("/tmp/a.py", loc=150, size=999, mtime="2025-01-01 00:00:00"),
    ]
    prev2 = read_previous_report(output)
    diff2 = compare_files(changed_files, prev2)
    updated2 = apply_audit_to_files(changed_files, diff2)

    assert updated2[0].audit_status == "Modificado"
    assert updated2[0].audit_marked == "No"


# ---------------------------------------------------------------------------
# Reporter audit columns
# ---------------------------------------------------------------------------

def test_csv_report_includes_audit_column(tmp_path):
    """CSV output contains ``Auditado`` and ``Estado`` columns with audit."""
    files = [_make_fileinfo("/tmp/a.py", loc=100)]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    diff = compare_files(files, {})
    updated = apply_audit_to_files(files, diff)
    audit_stats = calculate_audit_stats(updated, diff)

    output = str(tmp_path / "test.csv")
    CSVReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    content = open(output, "r", encoding="utf-8").read()
    assert "Auditado" in content
    assert "Estado" in content
    assert "Resumen de Auditoria" in content
    assert "Nuevo" in content


def test_csv_report_without_audit(tmp_path):
    """CSV output does NOT contain audit columns when audit_state=False."""
    files = [_make_fileinfo("/tmp/a.py", loc=100)]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    output = str(tmp_path / "test.csv")
    CSVReporter().generate(files, stats, output)

    content = open(output, "r", encoding="utf-8").read()
    assert "Auditado" not in content
    assert "Resumen de Auditoria" not in content


def test_xlsx_report_includes_audit_columns(tmp_path):
    """XLSX output contains audit columns and data validation when audit=True."""
    from openpyxl import load_workbook

    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00")]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    diff = compare_files(files, {})
    updated = apply_audit_to_files(files, diff)
    audit_stats = calculate_audit_stats(updated, diff)

    output = str(tmp_path / "test.xlsx")
    XLSXReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    wb = load_workbook(output)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Auditado" in headers
    assert "Estado" in headers

    # Check data validation exists
    assert len(ws.data_validations.dataValidation) > 0
    dv = ws.data_validations.dataValidation[0]
    assert "Si" in dv.formula1
    assert "No" in dv.formula1

    wb.close()


def test_json_report_includes_audit(tmp_path):
    """JSON output includes ``audit_status`` and ``audit_marked`` per file."""
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00")]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    diff = compare_files(files, {})
    updated = apply_audit_to_files(files, diff)
    audit_stats = calculate_audit_stats(updated, diff)

    output = str(tmp_path / "test.json")
    JSONReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    data = json.loads(open(output, "r", encoding="utf-8").read())
    assert "audit_status" in data["files"][0]
    assert "audit_marked" in data["files"][0]
    assert "audit_summary" in data
    assert data["audit_summary"]["new_files"] == 1


def test_json_report_without_audit(tmp_path):
    """JSON output does NOT include audit fields when audit_state=False."""
    files = [_make_fileinfo("/tmp/a.py", loc=100)]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    output = str(tmp_path / "test.json")
    JSONReporter().generate(files, stats, output)

    data = json.loads(open(output, "r", encoding="utf-8").read())
    assert "audit_summary" not in data


def test_html_report_includes_audit(tmp_path):
    """HTML output contains audit checkbox and progress when audit=True."""
    files = [_make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00")]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    diff = compare_files(files, {})
    updated = apply_audit_to_files(files, diff)
    audit_stats = calculate_audit_stats(updated, diff)

    output = str(tmp_path / "test.html")
    HTMLReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    content = open(output, "r", encoding="utf-8").read()
    assert "audit-checkbox" in content
    assert "var AUDIT_ENABLED=true" in content
    assert "Progreso de Auditoria" in content
    assert "estado-badge" in content


def test_html_report_without_audit(tmp_path):
    """HTML output does NOT contain audit elements when no audit."""
    files = [_make_fileinfo("/tmp/a.py", loc=100)]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    output = str(tmp_path / "test.html")
    HTMLReporter().generate(files, stats, output)

    content = open(output, "r", encoding="utf-8").read()
    assert "Progreso de Auditoria" not in content
    assert 'class="audit-checkbox"' not in content
    assert "var AUDIT_ENABLED=false" in content


# ---------------------------------------------------------------------------
# Backward compatibility: legacy AuditState still works
# ---------------------------------------------------------------------------

def test_audit_state_creation():
    """An empty ``AuditState`` has no entries."""
    state = AuditState()
    assert state.entries == {}
    assert state.last_updated == ""


def test_audit_state_mark():
    """``mark`` sets audited=True and populates metadata."""
    state = AuditState()
    state.mark("/tmp/foo.py", audited=True, auditor="Alice")
    entry = state.entries["/tmp/foo.py"]
    assert entry.audited is True
    assert entry.auditor == "Alice"
    assert entry.audit_date != ""


def test_load_save_roundtrip(tmp_path):
    """Save then load should preserve all data (legacy)."""
    state = AuditState()
    state.mark("/tmp/foo.py", audited=True, auditor="Alice", notes="Clean")
    state.mark("/tmp/bar.py", audited=False, auditor="Bob")

    filepath = str(tmp_path / "audit.json")
    save_audit_state(state, filepath)

    loaded = load_audit_state(filepath)
    assert loaded.is_audited("/tmp/foo.py") is True
    assert loaded.is_audited("/tmp/bar.py") is False


def test_merge_adds_missing_files():
    """``merge_audit_state`` adds scanned files (legacy)."""
    state = AuditState()
    files = [_make_fileinfo("/tmp/a.py", loc=50), _make_fileinfo("/tmp/b.py", loc=100)]
    state.mark("/tmp/a.py", audited=True, auditor="Alice")
    merged = merge_audit_state(state, files)
    assert merged.is_audited("/tmp/a.py") is True
    assert merged.is_audited("/tmp/b.py") is False


def test_legacy_stats():
    """Legacy ``AuditState.stats`` still works."""
    state = AuditState()
    files = [_make_fileinfo("/tmp/a.py", loc=100), _make_fileinfo("/tmp/b.py", loc=200)]
    state.mark("/tmp/a.py", audited=True)
    s = state.stats(files)
    assert s["audited_files"] == 1
    assert s["audited_loc"] == 100
    assert s["unaudited_loc"] == 200


# ---------------------------------------------------------------------------
# AuditDiff.determine_status
# ---------------------------------------------------------------------------

def test_audit_diff_determine_status():
    """``determine_status`` returns correct values."""
    diff = AuditDiff()
    diff.changes["/tmp/a.py"] = FileChangeStatus(status="Auditado", audited="Si")
    diff.changes["/tmp/b.py"] = FileChangeStatus(status="Modificado", audited="No")

    assert diff.determine_status("/tmp/a.py") == "Auditado"
    assert diff.determine_status("/tmp/b.py") == "Modificado"
    assert diff.determine_status("/tmp/unknown.py") == "Nuevo"


# ---------------------------------------------------------------------------
# Full real-world roundtrip: actual files → scan → XLSX → edit XLSX → re-scan
# ---------------------------------------------------------------------------

def test_full_roundtrip_real_files(tmp_path):
    """End-to-end roundtrip using real files, openpyxl edits, and re-scans.

    Covers: first scan → mark → second scan → modify file → third scan →
    add file → fourth scan.
    """
    import os
    import time
    from openpyxl import load_workbook
    from to_codigo.core.scanner import _collect_files, _process_file, DEFAULT_EXCLUDE_DIRS

    project = tmp_path / "project"
    project.mkdir()

    # Create 5 source files.
    for i in range(5):
        (project / f"file{i}.py").write_text(f"# file {i}\nprint({i})\n")

    output = str(tmp_path / "report.xlsx")

    def _scan():
        files = _collect_files(
            root=str(project),
            exclude_dirs=set(DEFAULT_EXCLUDE_DIRS),
            exclude_exts=set(),
            include_exts=set(),
            recursive=True,
            gitignore_patterns=[],
        )
        results = []
        for fpath in files:
            info, _ = _process_file((fpath, str(project)), collect_todos=False)
            if info:
                results.append(info)
        return results

    # --- First scan: all "Nuevo" ---
    files1 = _scan()
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files1)
    diff1 = compare_files(files1, {})
    updated1 = apply_audit_to_files(files1, diff1)
    audit_stats1 = calculate_audit_stats(updated1, diff1)
    XLSXReporter().generate(updated1, stats, output, audit_state=True, audit_stats=audit_stats1)

    assert diff1.new_files == 5
    assert diff1.audited_unchanged == 0

    # --- Mark first 3 files as "Si" in the XLSX ---
    wb = load_workbook(output)
    ws = wb.active
    for row in range(2, 5):
        ws.cell(row=row, column=13, value="Si")
    wb.save(output)
    wb.close()

    # --- Second scan: 3 audited, 2 pending ---
    files2 = _scan()
    prev2 = read_previous_report(output)
    assert len(prev2) == 5, f"Expected 5 entries in previous data, got {len(prev2)}"

    diff2 = compare_files(files2, prev2)
    updated2 = apply_audit_to_files(files2, diff2)

    assert diff2.audited_unchanged == 3, f"Expected 3 audited_unchanged, got {diff2.audited_unchanged}"
    assert diff2.pending_unchanged == 2
    assert diff2.new_files == 0
    assert diff2.audited_modified == 0

    audited_files = [f for f in updated2 if f.audit_status == "Auditado"]
    assert len(audited_files) == 3
    for f in audited_files:
        assert f.audit_marked == "Si"

    # --- Modify one of the AUDITED files ---
    time.sleep(1)
    target = audited_files[0]
    target_path = target.absolute_path
    with open(target_path, "w") as fh:
        fh.write("# modified\nprint('changed')\nx = 999\n")

    # --- Third scan: 2 audited, 1 modified, 2 pending ---
    files3 = _scan()
    prev3 = read_previous_report(output)
    diff3 = compare_files(files3, prev3)
    updated3 = apply_audit_to_files(files3, diff3)

    assert diff3.audited_unchanged == 2, f"Expected 2 audited_unchanged, got {diff3.audited_unchanged}"
    assert diff3.audited_modified >= 1, f"Expected at least 1 modified, got {diff3.audited_modified}"

    modified = [f for f in updated3 if f.audit_status == "Modificado"]
    assert len(modified) == 1
    assert modified[0].audit_marked == "No"
    assert modified[0].filename == target.filename

    # --- Add a new file ---
    (project / "new_file.py").write_text("z = 42\n")

    # --- Fourth scan: 2 audited, 1 modified, 2 pending, 1 new ---
    files4 = _scan()
    prev4 = read_previous_report(output)
    diff4 = compare_files(files4, prev4)
    updated4 = apply_audit_to_files(files4, diff4)

    assert diff4.audited_unchanged == 2
    assert diff4.new_files >= 1, f"Expected at least 1 new file, got {diff4.new_files}"

    new = [f for f in updated4 if f.audit_status == "Nuevo"]
    assert len(new) == 1
    assert new[0].filename == "new_file.py"


def test_relative_path_roundtrip(tmp_path):
    """Bug fix: scanning with a relative path (not starting with ./ or /).

    Previously, paths like ``myproject/file.py`` were stored in the XLSX but
    rejected by the reader's path-format filter, causing all marks to be lost.
    """
    import os
    from openpyxl import load_workbook
    from to_codigo.core.scanner import _collect_files, _process_file, DEFAULT_EXCLUDE_DIRS

    project = tmp_path / "relproj"
    project.mkdir()
    (project / "a.py").write_text("a = 1\n")
    (project / "b.py").write_text("b = 2\n")

    output = str(tmp_path / "relreport.xlsx")

    # Simulate scanning from within tmp_path using a relative directory name.
    orig_cwd = os.getcwd()
    try:
        os.chdir(str(tmp_path))
        files = _collect_files(
            root="relproj",
            exclude_dirs=set(DEFAULT_EXCLUDE_DIRS),
            exclude_exts=set(),
            include_exts=set(),
            recursive=True,
            gitignore_patterns=[],
        )
    finally:
        os.chdir(orig_cwd)

    assert len(files) == 2

    # All paths must be absolute (the bug was that they weren't).
    for fpath in files:
        assert os.path.isabs(fpath), f"Path should be absolute: {fpath}"

    # Generate XLSX.
    results = []
    for fpath in files:
        info, _ = _process_file((fpath, str(project)), collect_todos=False)
        if info:
            results.append(info)

    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(results)
    diff = compare_files(results, {})
    updated = apply_audit_to_files(results, diff)
    audit_stats = calculate_audit_stats(updated, diff)
    XLSXReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    # Mark all as "Si".
    wb = load_workbook(output)
    ws = wb.active
    for row in range(2, 4):
        ws.cell(row=row, column=13, value="Si")
    wb.save(output)
    wb.close()

    # Read back — should detect the marks.
    prev = read_previous_report(output)
    assert len(prev) == 2, f"Expected 2 entries, got {len(prev)}"
    for path, data in prev.items():
        assert data["audited"] == "Si", f"Expected 'Si' for {path}, got '{data['audited']}'"

    diff2 = compare_files(results, prev)
    assert diff2.audited_unchanged == 2, f"Expected 2 audited_unchanged, got {diff2.audited_unchanged}"


def test_xlsx_reader_skips_summary_rows(tmp_path):
    """The XLSX reader must not parse summary rows as file data."""
    from openpyxl import load_workbook

    files = [
        _make_fileinfo("/tmp/a.py", loc=100, size=100, mtime="2025-01-01 00:00:00"),
        _make_fileinfo("/tmp/b.py", loc=200, size=200, mtime="2025-01-01 00:00:00"),
    ]
    from to_codigo.core.scanner import compute_stats
    stats = compute_stats(files)

    diff = compare_files(files, {})
    updated = apply_audit_to_files(files, diff)
    audit_stats = calculate_audit_stats(updated, diff)

    output = str(tmp_path / "test.xlsx")
    XLSXReporter().generate(updated, stats, output, audit_state=True, audit_stats=audit_stats)

    result = read_previous_report(output)
    # Should only contain the 2 file paths, not summary rows.
    assert len(result) == 2
    assert "/tmp/a.py" in result
    assert "/tmp/b.py" in result
    # Summary markers should NOT appear as keys.
    assert not any(k.startswith("Resumen") for k in result)
    assert not any(k.startswith("===") for k in result)
    assert not any(k.startswith("Python") for k in result)
    assert not any(k.startswith("Archivos") for k in result)
