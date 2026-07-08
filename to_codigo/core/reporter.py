"""Report generation -- Strategy pattern.

Each reporter implements the :class:`Reporter` interface. Callers select a
concrete reporter at runtime via the ``REPORTERS`` registry:

.. code-block:: python

   reporter_cls = REPORTERS["xlsx"]
   reporter_cls().generate(files, stats, "output.xlsx")
"""

from __future__ import annotations

import csv
import json
import logging
import math
import os
from abc import ABC, abstractmethod
from dataclasses import asdict
from datetime import datetime
from html import escape
from typing import Any

from to_codigo.core.models import FileInfo, LanguageStats, TodoItem, AuditState

logger = logging.getLogger(__name__)

HEADERS: tuple[str, ...] = (
    "Ruta_Absoluta_Archivo",
    "Ruta_Relativa_Carpeta",
    "Nombre_Archivo",
    "Lenguaje_Programacion",
    "Tamano_Bytes",
    "Fecha_Modificacion",
    "Lineas_Totales",
    "Lineas_Codigo",
    "Lineas_Comentarios",
    "Lineas_Blanco",
    "TODOs",
    "FIXMEs",
)

SUMMARY_HEADERS: tuple[str, ...] = (
    "Lenguaje",
    "Archivos",
    "Lineas_Codigo_Total",
    "Lineas_Comentarios_Total",
    "Lineas_Blanco_Total",
    "Lineas_Totales",
    "TODOs",
    "FIXMEs",
)

AUDIT_HEADERS: tuple[str, ...] = (
    "Auditado",
    "Estado",
)

# Colour palette for charts (cycle through these).
LANG_COLORS: tuple[str, ...] = (
    "#7dcfff", "#9ece6a", "#ff9e64", "#f7768e",
    "#7aa2f7", "#bb9af7", "#e0af68", "#73daca",
    "#c0caf5", "#b4f9f8", "#2ac3de", "#89ddff",
)


def _file_to_row(info: FileInfo, audit: bool = False) -> list[Any]:
    """Convert a :class:`FileInfo` to a list matching :data:`HEADERS`.

    When *audit* is ``True``, appends the ``Auditado`` and ``Estado`` columns.
    """
    row = [
        info.absolute_path,
        info.relative_path,
        info.filename,
        info.language,
        info.size_bytes,
        info.modified_at,
        info.total_lines,
        info.code_lines,
        info.comment_lines,
        info.blank_lines,
        info.todos,
        info.fixmes,
    ]
    if audit:
        row.append(info.audit_marked)
        row.append(info.audit_status)
    return row


# ---------------------------------------------------------------------------
# Abstract strategy
# ---------------------------------------------------------------------------

class Reporter(ABC):
    """Abstract base class for all report strategies."""

    @abstractmethod
    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        """Write the report to *output_path*.

        Args:
            files: List of per-file records.
            stats: Per-language aggregate statistics.
            output_path: Destination file path.
            audit_state: When ``True``, audit columns and summary sections
                are included. FileInfo objects already carry audit data
                in ``audit_status`` and ``audit_marked`` fields.
            audit_stats: Optional pre-computed audit statistics dict.
        """
        ...


# ---------------------------------------------------------------------------
# CSV strategy
# ---------------------------------------------------------------------------

class CSVReporter(Reporter):
    """Comma-separated values report (RFC 4180 compliant via :mod:`csv`)."""

    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        audit = audit_state is not False
        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            if audit:
                writer.writerow(list(HEADERS) + list(AUDIT_HEADERS))
            else:
                writer.writerow(HEADERS)
            for info in files:
                writer.writerow(_file_to_row(info, audit=audit))

            writer.writerow([])
            writer.writerow(["Resumen por Lenguaje"])
            writer.writerow(SUMMARY_HEADERS)
            for lang, data in stats.items():
                writer.writerow([
                    lang,
                    data.files,
                    data.total_code_lines,
                    data.total_comment_lines,
                    data.total_blank_lines,
                    data.total_lines,
                    data.total_todos,
                    data.total_fixmes,
                ])

            if audit and audit_stats:
                a = audit_stats
                writer.writerow([])
                writer.writerow(["=== Resumen de Auditoria ==="])
                writer.writerow(["Archivos Auditados (sin cambios)", a["audited_files"]])
                writer.writerow(["Lineas Auditadas", a["audited_loc"]])
                writer.writerow(["Archivos Modificados (re-auditar)", a["modified_files"]])
                writer.writerow(["Lineas Modificadas", a["modified_loc"]])
                writer.writerow(["Archivos Nuevos", a["new_files"]])
                writer.writerow(["Lineas Nuevas", a["new_loc"]])
                writer.writerow(["Archivos Pendientes", a["pending_files"]])
                writer.writerow(["Lineas Pendientes", a["pending_loc"]])
                writer.writerow(["Archivos Eliminados", a.get("removed_files", 0)])
                writer.writerow(["Archivos Totales", a["total_files"]])
                writer.writerow(["Lineas Totales", a["total_loc"]])
                writer.writerow(["% Auditado", f'{a["pct_audited"]:.1f}%'])
        logger.info("CSV report generated: %s", output_path)


# ---------------------------------------------------------------------------
# XLSX strategy
# ---------------------------------------------------------------------------

class XLSXReporter(Reporter):
    """Excel report with styled headers, auto-filter, frozen panes, and
    a colour-coded summary section.

    When audit is enabled:
    - ``Auditado`` column has a data-validation dropdown (Si/No).
    - ``Estado`` column shows computed status with row-level conditional formatting.
    - Audit summary section appears below the per-language summary.

    Requires the ``openpyxl`` package.
    """

    # Reasonable column widths (in Excel units).
    _COLUMN_WIDTHS: tuple[int, ...] = (
        60,  # absolute path
        35,  # relative path
        30,  # filename
        18,  # language
        14,  # size
        22,  # modified at
        14,  # total lines
        14,  # code lines
        16,  # comment lines
        14,  # blank lines
        10,  # TODOs
        10,  # FIXMEs
    )
    _AUDIT_COLUMN_WIDTHS: tuple[int, ...] = (12, 16)

    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        audit = audit_state is not False
        wb = Workbook()
        ws = wb.active
        ws.title = "to-codigo"

        # --- Styles ---
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        audit_header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        summary_header_font = Font(bold=True, color="FFFFFF", size=11)
        summary_header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        summary_title_font = Font(bold=True, size=13, color="C55A11")

        # --- Determine headers ---
        all_headers = list(HEADERS)
        if audit:
            all_headers += list(AUDIT_HEADERS)

        # --- Header row ---
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            if audit and col > len(HEADERS):
                cell.fill = audit_header_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Column widths ---
        widths = list(self._COLUMN_WIDTHS)
        if audit:
            widths += list(self._AUDIT_COLUMN_WIDTHS)
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # --- Data rows ---
        for row_idx, info in enumerate(files, 2):
            row_data = _file_to_row(info, audit=audit)
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        last_data_row = len(files) + 1

        # --- Audit data validation + conditional formatting ---
        if audit and files:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.formatting.rule import FormulaRule

            audit_col = len(HEADERS) + 1  # "Auditado"
            estado_col = len(HEADERS) + 2  # "Estado"
            audit_letter = get_column_letter(audit_col)
            estado_letter = get_column_letter(estado_col)
            last_letter = get_column_letter(len(all_headers))

            # --- Data validation dropdown for Auditado column ---
            dv = DataValidation(
                type="list",
                formula1='"Si,No"',
                allow_blank=True,
            )
            dv.add(f"{audit_letter}2:{audit_letter}{last_data_row}")
            ws.add_data_validation(dv)

            # --- Conditional formatting by Estado column ---
            # "Auditado" → light green row
            green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            ws.conditional_formatting.add(
                f"A2:{last_letter}{last_data_row}",
                FormulaRule(
                    formula=[f'${estado_letter}2="Auditado"'],
                    stopIfTrue=False,
                    fill=green_fill,
                ),
            )

            # "Modificado" → light yellow row (WARNING)
            yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            ws.conditional_formatting.add(
                f"A2:{last_letter}{last_data_row}",
                FormulaRule(
                    formula=[f'${estado_letter}2="Modificado"'],
                    stopIfTrue=False,
                    fill=yellow_fill,
                ),
            )

            # "Nuevo" → light blue row
            blue_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
            ws.conditional_formatting.add(
                f"A2:{last_letter}{last_data_row}",
                FormulaRule(
                    formula=[f'${estado_letter}2="Nuevo"'],
                    stopIfTrue=False,
                    fill=blue_fill,
                ),
            )

        # --- Auto-filter on header row ---
        if files:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{last_data_row}"

        # --- Freeze top row ---
        ws.freeze_panes = "A2"

        # --- Summary section ---
        summary_start = last_data_row + 3  # gap row

        title_cell = ws.cell(row=summary_start, column=1, value="Resumen por Lenguaje")
        title_cell.font = summary_title_font

        summary_header_row = summary_start + 1
        for col, header in enumerate(SUMMARY_HEADERS, 1):
            cell = ws.cell(row=summary_header_row, column=col, value=header)
            cell.font = summary_header_font
            cell.fill = summary_header_fill
            cell.alignment = Alignment(horizontal="center")

        for offset, (lang, data) in enumerate(stats.items()):
            row = summary_header_row + 1 + offset
            ws.cell(row=row, column=1, value=lang)
            ws.cell(row=row, column=2, value=data.files)
            ws.cell(row=row, column=3, value=data.total_code_lines)
            ws.cell(row=row, column=4, value=data.total_comment_lines)
            ws.cell(row=row, column=5, value=data.total_blank_lines)
            ws.cell(row=row, column=6, value=data.total_lines)
            ws.cell(row=row, column=7, value=data.total_todos)
            ws.cell(row=row, column=8, value=data.total_fixmes)

        # --- Audit summary section ---
        if audit and audit_stats:
            a = audit_stats
            audit_start = summary_header_row + len(stats) + 3
            audit_title = ws.cell(row=audit_start, column=1, value="=== Resumen de Auditoria ===")
            audit_title.font = Font(bold=True, size=13, color="2D6A4F")

            audit_rows = [
                ("Archivos Auditados (sin cambios)", a["audited_files"], f'{a["audited_loc"]:,} LOC'),
                ("Archivos Modificados (re-auditar)", a["modified_files"], f'{a["modified_loc"]:,} LOC'),
                ("Archivos Nuevos", a["new_files"], f'{a["new_loc"]:,} LOC'),
                ("Archivos Pendientes", a["pending_files"], f'{a["pending_loc"]:,} LOC'),
                ("Archivos Eliminados", a.get("removed_files", 0), ""),
                ("TOTAL", a["total_files"], f'{a["total_loc"]:,} LOC'),
                ("% Auditado", f'{a["pct_audited"]:.1f}%', ""),
            ]
            for i, (label, value, extra) in enumerate(audit_rows, 1):
                ws.cell(row=audit_start + i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=audit_start + i, column=2, value=value)
                if extra:
                    ws.cell(row=audit_start + i, column=3, value=extra)

        wb.save(output_path)
        logger.info("XLSX report generated: %s", output_path)


# ---------------------------------------------------------------------------
# JSON strategy
# ---------------------------------------------------------------------------

class JSONReporter(Reporter):
    """Structured JSON report with file details and per-language summary."""

    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        audit = audit_state is not False
        file_list: list[dict[str, Any]] = []
        for f in files:
            entry_dict = asdict(f)
            if audit:
                entry_dict["audit_status"] = f.audit_status
                entry_dict["audit_marked"] = f.audit_marked
            file_list.append(entry_dict)

        payload: dict[str, Any] = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "files": file_list,
            "summary": {
                lang: {
                    "files": s.files,
                    "total_code_lines": s.total_code_lines,
                    "total_comment_lines": s.total_comment_lines,
                    "total_blank_lines": s.total_blank_lines,
                    "total_lines": s.total_lines,
                    "total_todos": s.total_todos,
                    "total_fixmes": s.total_fixmes,
                    "total_hacks": s.total_hacks,
                    "total_notes": s.total_notes,
                }
                for lang, s in stats.items()
            },
        }
        if audit and audit_stats:
            payload["audit_summary"] = audit_stats
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        logger.info("JSON report generated: %s", output_path)


# ---------------------------------------------------------------------------
# Markdown strategy
# ---------------------------------------------------------------------------

class MarkdownReporter(Reporter):
    """Markdown report with summary tables, top files, and tech debt section."""

    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        from to_codigo.assets.banner import BANNER, VERSION

        audit = audit_state is not False
        total_code = sum(s.total_code_lines for s in stats.values())
        total_comment = sum(s.total_comment_lines for s in stats.values())
        total_blank = sum(s.total_blank_lines for s in stats.values())
        total_lines = sum(s.total_lines for s in stats.values())
        total_todos = sum(s.total_todos for s in stats.values())
        total_fixmes = sum(s.total_fixmes for s in stats.values())

        lines: list[str] = []

        lines.append("# to-codigo Report\n")
        lines.append("```\n" + BANNER.strip("\n") + "\n```\n")
        lines.append(f"**v{VERSION}** - Code Analyzer & Inventory Tool\n")
        lines.append("---\n")

        # Summary table
        lines.append("## Summary\n")
        lines.append("| Metric | Value |")
        lines.append("|--------|-------|")
        lines.append(f"| Total Files | {len(files)} |")
        lines.append(f"| Total LOC | {total_code:,} |")
        lines.append(f"| Total Comments | {total_comment:,} |")
        lines.append(f"| Total Blank | {total_blank:,} |")
        lines.append(f"| Total Lines | {total_lines:,} |")
        lines.append(f"| TODOs | {total_todos} |")
        lines.append(f"| FIXMEs | {total_fixmes} |")
        lines.append("")

        # Per-language table
        audit_col = " | Auditado" if audit else ""
        audit_sep = "|----------" if audit else ""
        lines.append("## Per-Language Breakdown\n")
        lines.append(f"| Language | Files | LOC | Comments | Blank | Total | TODOs | FIXMEs{audit_col} |")
        lines.append(f"|----------|-------|-----|----------|-------|-------|-------|--------{audit_sep}|")
        sorted_stats = sorted(stats.items(), key=lambda kv: kv[1].total_code_lines, reverse=True)
        for lang, data in sorted_stats:
            audit_mark = ""
            if audit:
                audited_count = sum(
                    1 for f in files
                    if f.language == lang and f.audit_marked == "Si"
                )
                audit_mark = f" | {audited_count}/{data.files}"
            lines.append(
                f"| {lang} | {data.files} | {data.total_code_lines:,} | "
                f"{data.total_comment_lines:,} | {data.total_blank_lines:,} | "
                f"{data.total_lines:,} | {data.total_todos} | {data.total_fixmes}{audit_mark} |"
            )
        lines.append("")

        # Top 10 files
        top_files = sorted(files, key=lambda f: f.code_lines, reverse=True)[:10]
        if top_files:
            lines.append("## Top 10 Files by LOC\n")
            top_audit_col = " | Auditado | Estado" if audit else ""
            lines.append(f"| Rank | File | Language | LOC{top_audit_col} |")
            lines.append(f"|------|------|----------|-----{'|----------|--------' if audit else ''}|")
            for i, info in enumerate(top_files, 1):
                mark = ""
                if audit:
                    symbol = "✓" if info.audit_marked == "Si" else "✗"
                    mark = f" | {symbol} | {info.audit_status}"
                lines.append(f"| {i} | {info.filename} | {info.language} | {info.code_lines:,}{mark} |")
            lines.append("")

        # Tech debt
        debt_files = [f for f in files if f.todos > 0 or f.fixmes > 0 or f.hacks > 0]
        if debt_files:
            lines.append("## Tech Debt\n")
            lines.append("| File | TODOs | FIXMEs | HACKs |")
            lines.append("|------|-------|--------|-------|")
            for info in debt_files:
                lines.append(
                    f"| {info.filename} | {info.todos} | {info.fixmes} | {info.hacks} |"
                )
            lines.append("")

        # Audit summary
        if audit and audit_stats:
            a = audit_stats
            lines.append("## Audit Summary\n")
            lines.append("| Metric | Value | LOC |")
            lines.append("|--------|-------|-----|")
            lines.append(f"| Auditados (sin cambios) | {a['audited_files']} | {a['audited_loc']:,} |")
            lines.append(f"| Modificados (re-auditar) | {a['modified_files']} | {a['modified_loc']:,} |")
            lines.append(f"| Nuevos | {a['new_files']} | {a['new_loc']:,} |")
            lines.append(f"| Pendientes | {a['pending_files']} | {a['pending_loc']:,} |")
            lines.append(f"| Eliminados | {a.get('removed_files', 0)} | — |")
            lines.append(f"| **TOTAL** | **{a['total_files']}** | **{a['total_loc']:,}** |")
            lines.append(f"| % Auditado | {a['pct_audited']:.1f}% | — |")
            lines.append("")

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines.append(f"\n---\n*Generated: {timestamp}*\n")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        logger.info("Markdown report generated: %s", output_path)


# ---------------------------------------------------------------------------
# HTML strategy
# ---------------------------------------------------------------------------

class HTMLReporter(Reporter):
    """Self-contained HTML dashboard with charts, tables, and tech debt.

    Reads a template file, fills in placeholders with real data, and writes
    the output. Fully offline (no external CDNs or resources).
    """

    _TEMPLATE_PATH = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "templates",
        "dashboard.html",
    )

    def generate(
        self,
        files: list[FileInfo],
        stats: dict[str, LanguageStats],
        output_path: str,
        audit_state: bool = False,
        audit_stats: dict[str, Any] | None = None,
    ) -> None:
        from to_codigo.assets.banner import BANNER, VERSION

        audit = audit_state is not False

        with open(self._TEMPLATE_PATH, "r", encoding="utf-8") as f:
            template = f.read()

        # --- Summary numbers ---
        total_code = sum(s.total_code_lines for s in stats.values())
        total_comment = sum(s.total_comment_lines for s in stats.values())

        # --- Sorted stats ---
        sorted_stats = sorted(
            stats.items(),
            key=lambda kv: kv[1].total_code_lines,
            reverse=True,
        )

        # --- Bar chart ---
        max_loc = max(
            (s.total_code_lines for _, s in sorted_stats),
            default=1,
        )
        if max_loc == 0:
            max_loc = 1

        bar_parts: list[str] = []
        for idx, (lang, data) in enumerate(sorted_stats):
            pct = (data.total_code_lines / max_loc) * 100
            color = LANG_COLORS[idx % len(LANG_COLORS)]
            bar_parts.append(
                f'<div class="bar-row">'
                f'<div class="bar-label">{escape(lang)}</div>'
                f'<div class="bar-track">'
                f'<div class="bar-fill" style="width:{pct:.1f}%;background:{color}"></div>'
                f'<div class="bar-value">{data.total_code_lines:,}</div>'
                f'</div></div>'
            )
        bar_html = "\n".join(bar_parts)

        # --- Pie chart (SVG) ---
        total_files = sum(s.files for _, s in sorted_stats)
        pie_svg = self._build_pie_svg(sorted_stats, total_files)
        pie_legend = self._build_pie_legend(sorted_stats, total_files)

        # --- Audit data ---
        audit_progress_html = ""
        audit_filters_html = ""
        audit_table_headers = ""
        audit_data_json = "[]"

        if audit and audit_stats:
            a = audit_stats
            pct_loc = a["pct_audited"]

            # Progress bar + cards
            audit_progress_html = (
                f'<div class="section" id="audit-section">'
                f"<h2>Progreso de Auditoria</h2>"
                f'<div class="audit-progress-wrapper">'
                f'<div class="audit-progress-bar">'
                f'<div class="audit-progress-fill" id="audit-progress-fill" '
                f'style="width:{pct_loc:.1f}%"></div>'
                f'</div>'
                f'<div class="audit-progress-text" id="audit-progress-text">'
                f'{a["audited_loc"]:,} / {a["total_loc"]:,} lineas auditadas '
                f'({pct_loc:.1f}%)'
                f'</div>'
                f'</div>'
                f'<div class="cards audit-cards">'
                f'<div class="card card-green"><div class="label">Auditados</div>'
                f'<div class="value" id="audit-files-count">{a["audited_files"]}</div></div>'
                f'<div class="card card-orange"><div class="label">Modificados</div>'
                f'<div class="value" id="modified-files-count">{a["modified_files"]}</div></div>'
                f'<div class="card card-cyan"><div class="label">Nuevos</div>'
                f'<div class="value" id="new-files-count">{a["new_files"]}</div></div>'
                f'<div class="card card-red"><div class="label">Pendientes</div>'
                f'<div class="value" id="pending-files-count">{a["pending_files"]}</div></div>'
                f'<div class="card card-cyan"><div class="label">% Progreso</div>'
                f'<div class="value" id="audit-pct">{pct_loc:.1f}%</div></div>'
                f"</div>"
                f'<div class="audit-saved" id="audit-saved">Guardado &#10003;</div>'
                f"</div>"
            )

            # Filters + export/import
            audit_filters_html = (
                f'<div class="audit-toolbar">'
                f'<button class="audit-btn active" data-filter="all">Todos</button>'
                f'<button class="audit-btn" data-filter="auditado">Auditados &#10003;</button>'
                f'<button class="audit-btn" data-filter="modificado">Modificados &#9888;</button>'
                f'<button class="audit-btn" data-filter="nuevo">Nuevos +</button>'
                f'<button class="audit-btn" data-filter="pendiente">Pendientes</button>'
                f'<span class="audit-sep"></span>'
                f'<button class="audit-btn" id="audit-export">Exportar a Excel</button>'
                f'<input type="file" id="audit-import-input" accept=".json" style="display:none">'
                f"</div>"
            )

            # Extra table headers
            audit_table_headers = (
                "<th>Auditado</th>"
                "<th>Estado</th>"
            )

            # JSON data for JS
            audit_entries: list[dict] = []
            for info in files:
                audit_entries.append({
                    "path": info.absolute_path,
                    "audited": info.audit_marked == "Si",
                    "audit_marked": info.audit_marked,
                    "estado": info.audit_status,
                    "code_lines": info.code_lines,
                    "filename": info.filename,
                    "language": info.language,
                })
            audit_data_json = json.dumps(audit_entries, ensure_ascii=False)

        # --- Table rows ---
        table_rows: list[str] = []
        for info in sorted(files, key=lambda f: f.code_lines, reverse=True):
            row_parts: list[str] = ["<tr>"]
            if audit:
                checked = "checked" if info.audit_marked == "Si" else ""
                estado_class = f'audit-estado-{info.audit_status.lower()}'
                row_parts.append(
                    f'<td class="audit-col"><input type="checkbox" class="audit-checkbox" '
                    f'data-path="{escape(info.absolute_path)}" {checked}></td>'
                )
                row_parts.append(
                    f'<td class="audit-col estado-col"><span class="estado-badge {estado_class}">'
                    f'{escape(info.audit_status)}</span></td>'
                )
            row_parts.append(f'<td data-val="{escape(info.filename)}">{escape(info.filename)}</td>')
            row_parts.append(
                f'<td data-val="{escape(info.language)}"><span class="lang-badge" '
                f'style="background:#2a2e3f;color:#7dcfff">{escape(info.language)}</span></td>'
            )
            row_parts.append(f'<td data-val="{info.code_lines}">{info.code_lines:,}</td>')
            row_parts.append(f'<td data-val="{info.comment_lines}">{info.comment_lines:,}</td>')
            row_parts.append(f'<td data-val="{info.blank_lines}">{info.blank_lines:,}</td>')
            row_parts.append(f'<td data-val="{info.total_lines}">{info.total_lines:,}</td>')
            row_parts.append("</tr>")
            table_rows.append("".join(row_parts))
        table_html = "\n".join(table_rows)

        # --- Top 10 files ---
        top_files = sorted(files, key=lambda f: f.code_lines, reverse=True)[:10]
        top_parts: list[str] = []
        for i, info in enumerate(top_files, 1):
            top_parts.append(
                f'<div class="top-file">'
                f'<div class="top-rank">{i}</div>'
                f'<div class="top-info">'
                f'<div class="top-name">{escape(info.filename)}</div>'
                f'<div class="top-lang">{escape(info.language)} - {escape(info.relative_path)}</div>'
                f"</div>"
                f'<div class="top-loc">{info.code_lines:,} LOC</div>'
                f"</div>"
            )
        top_html = "\n".join(top_parts) if top_parts else '<div class="no-data">No files found</div>'

        # --- Tech debt ---
        debt_parts: list[str] = []
        debt_files = sorted(
            [f for f in files if f.todos + f.fixmes + f.hacks + f.notes > 0],
            key=lambda f: f.todos + f.fixmes + f.hacks + f.notes,
            reverse=True,
        )
        if debt_files:
            for info in debt_files:
                parts: list[str] = []
                if info.todos:
                    parts.append(f'<span class="debt-marker" style="background:#ff9e64">TODO: {info.todos}</span>')
                if info.fixmes:
                    parts.append(f'<span class="debt-marker" style="background:#f7768e">FIXME: {info.fixmes}</span>')
                if info.hacks:
                    parts.append(f'<span class="debt-marker" style="background:#e0af68">HACK: {info.hacks}</span>')
                if info.notes:
                    parts.append(f'<span class="debt-marker" style="background:#7aa2f7">NOTE: {info.notes}</span>')
                debt_parts.append(
                    f'<div class="debt-item">'
                    f'<div class="debt-header">'
                    f"{' '.join(parts)}"
                    f'<span class="debt-file">{escape(info.filename)}</span>'
                    f'<span class="debt-line">({info.relative_path})</span>'
                    f"</div>"
                    f"</div>"
                )
        debt_html = "\n".join(debt_parts) if debt_parts else '<div class="no-data">No TODOs, FIXMEs, or HACKs found. Clean codebase!</div>'

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- Fill template ---
        result = template
        replacements = {
            "{{BANNER}}": escape(BANNER.strip("\n")),
            "{{VERSION}}": VERSION,
            "{{TOTAL_FILES}}": str(len(files)),
            "{{TOTAL_LOC}}": f"{total_code:,}",
            "{{TOTAL_COMMENTS}}": f"{total_comment:,}",
            "{{LANGUAGES_COUNT}}": str(len(stats)),
            "{{BAR_CHART_HTML}}": bar_html,
            "{{PIE_CHART_SVG}}": pie_svg,
            "{{PIE_LEGEND_HTML}}": pie_legend,
            "{{TABLE_ROWS_HTML}}": table_html,
            "{{TOP_FILES_HTML}}": top_html,
            "{{TECH_DEBT_HTML}}": debt_html,
            "{{TIMESTAMP}}": timestamp,
            "{{AUDIT_ENABLED}}": "true" if audit else "false",
            "{{AUDIT_PROGRESS_HTML}}": audit_progress_html,
            "{{AUDIT_FILTERS_HTML}}": audit_filters_html,
            "{{AUDIT_TABLE_HEADERS}}": audit_table_headers,
            "{{AUDIT_DATA_JSON}}": audit_data_json,
        }
        for placeholder, value in replacements.items():
            result = result.replace(placeholder, value)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(result)
        logger.info("HTML report generated: %s", output_path)

    @staticmethod
    def _build_pie_svg(
        sorted_stats: list[tuple[str, LanguageStats]],
        total_files: int,
    ) -> str:
        """Build an inline SVG pie chart with stroke-dasharray segments."""
        radius = 80
        circumference = 2 * math.pi * radius
        offset = 0

        segments: list[str] = []
        for idx, (lang, data) in enumerate(sorted_stats):
            if total_files == 0:
                break
            fraction = data.files / total_files
            length = fraction * circumference
            color = LANG_COLORS[idx % len(LANG_COLORS)]

            if length < 0.5:
                continue

            segments.append(
                f'<circle cx="100" cy="100" r="{radius}" fill="none" '
                f'stroke="{color}" stroke-width="30" '
                f'stroke-dasharray="{length:.2f} {circumference - length:.2f}" '
                f'stroke-dashoffset="{-offset:.2f}" '
                f'transform="rotate(-90 100 100)" />'
            )
            offset += length

        svg = (
            f'<svg width="200" height="200" viewBox="0 0 200 200" '
            f'xmlns="http://www.w3.org/2000/svg">'
        )
        # Background circle
        svg += (
            f'<circle cx="100" cy="100" r="{radius}" fill="none" '
            f'stroke="#2a2e3f" stroke-width="30" />'
        )
        svg += "".join(segments)
        svg += f'<text x="100" y="95" text-anchor="middle" fill="#a9b1d6" font-size="28" font-weight="bold">{total_files}</text>'
        svg += '<text x="100" y="115" text-anchor="middle" fill="#565f89" font-size="12">files</text>'
        svg += "</svg>"
        return svg

    @staticmethod
    def _build_pie_legend(
        sorted_stats: list[tuple[str, LanguageStats]],
        total_files: int,
    ) -> str:
        """Build the legend HTML for the pie chart."""
        parts: list[str] = []
        for idx, (lang, data) in enumerate(sorted_stats):
            color = LANG_COLORS[idx % len(LANG_COLORS)]
            pct = (data.files / total_files * 100) if total_files > 0 else 0
            parts.append(
                f'<div class="legend-item">'
                f'<div class="legend-color" style="background:{color}"></div>'
                f'<span class="legend-label">{escape(lang)}</span>'
                f'<span class="legend-value">{data.files} ({pct:.1f}%)</span>'
                f"</div>"
            )
        return "\n".join(parts)


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------

REPORTERS: dict[str, type[Reporter]] = {
    "csv": CSVReporter,
    "xlsx": XLSXReporter,
    "json": JSONReporter,
    "html": HTMLReporter,
    "md": MarkdownReporter,
}
