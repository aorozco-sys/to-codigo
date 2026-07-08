"""Audit tracking — Excel/CSV/JSON report as the source of truth.

The previous implementation used a separate JSON state file (``.to-codigo-audit.json``).
The new design treats the **output report itself** as the audit state.

Workflow:
1. First run generates an Excel with ``Auditado`` (all ``"No"``) and ``Estado`` columns.
2. The auditor opens the Excel, marks ``Auditado = "Si"`` for reviewed files.
3. Second run detects the previous report, compares file size + mtime,
   and generates a new report reflecting changes.

Key functions:
- :func:`read_previous_report` — extract audit data from a previous XLSX/CSV/JSON.
- :func:`compare_files` — diff current scan vs. previous report data.
- :func:`calculate_audit_stats` — compute audit statistics (counts + LOC).
- :func:`apply_audit_to_files` — update each ``FileInfo`` with audit status and mark.

Backward-compat functions (:func:`load_audit_state`, :func:`save_audit_state`,
:func:`mark_file`, :func:`merge_audit_state`) are retained so that existing
code paths that still import them don't break, but the new ``--audit`` flow
no longer uses them.
"""

from __future__ import annotations

import csv
import json
import logging
import os
from datetime import datetime
from typing import Any

from to_codigo.core.models import (
    AuditDiff,
    AuditEntry,
    AuditState,
    FileInfo,
    FileChangeStatus,
)

logger = logging.getLogger(__name__)

# Column name constants matching reporter HEADERS.
_PATH_COL = "Ruta_Absoluta_Archivo"
_SIZE_COL = "Tamano_Bytes"
_MTIME_COL = "Fecha_Modificacion"
_AUDITED_COL = "Auditado"
_AUDITOR_COL = "Auditor"
_AUDIT_DATE_COL = "Fecha_Auditoria"


# ---------------------------------------------------------------------------
# New API — report-as-source-of-truth
# ---------------------------------------------------------------------------

def read_previous_report(report_path: str) -> dict[str, dict[str, Any]]:
    """Read a previous Excel/CSV/JSON report and extract audit data.

    Returns a mapping::

        {filepath: {"audited": "Si"/"No", "size": int, "mtime": str, "auditor": str}}

    Handles:
    - **.xlsx**: Uses ``openpyxl`` to read columns.
    - **.csv**: Parses CSV rows for the same columns.
    - **.json**: Parses JSON file objects.
    - If the file doesn't exist or can't be read: returns ``{}``.
    """
    if not os.path.isfile(report_path):
        return {}

    ext = os.path.splitext(report_path)[1].lower()
    try:
        if ext == ".xlsx":
            return _read_xlsx_report(report_path)
        elif ext == ".csv":
            return _read_csv_report(report_path)
        elif ext == ".json":
            return _read_json_report(report_path)
        else:
            logger.debug("Unsupported report format for audit reading: %s", ext)
            return {}
    except Exception as e:
        logger.warning("Could not read previous report %s: %s", report_path, e)
        return {}


def _read_xlsx_report(path: str) -> dict[str, dict[str, Any]]:
    """Extract audit data from an XLSX report."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    # Read header row to find column indices.
    try:
        headers = list(next(rows))
    except StopIteration:
        wb.close()
        return {}

    # Build a header -> index map (strip whitespace).
    hmap: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h is not None:
            hmap[str(h).strip()] = idx

    if _PATH_COL not in hmap:
        wb.close()
        return {}

    path_idx = hmap[_PATH_COL]
    audited_idx = hmap.get(_AUDITED_COL)
    size_idx = hmap.get(_SIZE_COL)
    mtime_idx = hmap.get(_MTIME_COL)
    auditor_idx = hmap.get(_AUDITOR_COL)

    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row is None:
            continue
        fpath = row[path_idx] if path_idx < len(row) else None
        if not fpath:
            continue
        fpath = str(fpath).strip()
        if not fpath or fpath.startswith("Resumen") or fpath.startswith("==="):
            continue
        # Skip summary rows — only accept paths that look like absolute file paths.
        if not (fpath.startswith("/") or fpath[1:3] == ":\\" or fpath.startswith("./") or fpath.startswith("../")):
            continue

        entry: dict[str, Any] = {
            "audited": "No",
            "size": 0,
            "mtime": "",
            "auditor": "",
        }
        if audited_idx is not None and audited_idx < len(row):
            val = row[audited_idx]
            entry["audited"] = "Si" if str(val).strip() in ("Si", "si", "SI", "True", "true", "1") else "No"
        if size_idx is not None and size_idx < len(row):
            try:
                entry["size"] = int(row[size_idx]) if row[size_idx] else 0
            except (ValueError, TypeError):
                entry["size"] = 0
        if mtime_idx is not None and mtime_idx < len(row):
            entry["mtime"] = str(row[mtime_idx]).strip() if row[mtime_idx] else ""
        if auditor_idx is not None and auditor_idx < len(row):
            entry["auditor"] = str(row[auditor_idx]).strip() if row[auditor_idx] else ""

        result[fpath] = entry

    wb.close()
    return result


def _read_csv_report(path: str) -> dict[str, dict[str, Any]]:
    """Extract audit data from a CSV report."""
    result: dict[str, dict[str, Any]] = {}
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return {}

        headers = [h.strip() for h in headers]
        hmap = {h: idx for idx, h in enumerate(headers) if h}

        if _PATH_COL not in hmap:
            return {}

        path_idx = hmap[_PATH_COL]
        audited_idx = hmap.get(_AUDITED_COL)
        size_idx = hmap.get(_SIZE_COL)
        mtime_idx = hmap.get(_MTIME_COL)
        auditor_idx = hmap.get(_AUDITOR_COL)

        for row in reader:
            if not row or len(row) <= path_idx:
                continue
            fpath = row[path_idx].strip()
            if not fpath or fpath.startswith("Resumen") or fpath.startswith("==="):
                continue
            # Skip summary rows — only accept paths that look like absolute file paths.
            if not (fpath.startswith("/") or fpath[1:3] == ":\\" or fpath.startswith("./") or fpath.startswith("../")):
                continue

            entry: dict[str, Any] = {
                "audited": "No",
                "size": 0,
                "mtime": "",
                "auditor": "",
            }
            if audited_idx is not None and audited_idx < len(row):
                val = row[audited_idx].strip()
                entry["audited"] = "Si" if val in ("Si", "si", "SI", "True", "true", "1") else "No"
            if size_idx is not None and size_idx < len(row):
                try:
                    entry["size"] = int(row[size_idx]) if row[size_idx] else 0
                except (ValueError, TypeError):
                    entry["size"] = 0
            if mtime_idx is not None and mtime_idx < len(row):
                entry["mtime"] = row[mtime_idx].strip()
            if auditor_idx is not None and auditor_idx < len(row):
                entry["auditor"] = row[auditor_idx].strip()

            result[fpath] = entry

    return result


def _read_json_report(path: str) -> dict[str, dict[str, Any]]:
    """Extract audit data from a JSON report."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    result: dict[str, dict[str, Any]] = {}
    for fobj in data.get("files", []):
        abs_path = fobj.get("absolute_path") or fobj.get("Ruta_Absoluta_Archivo") or ""
        if not abs_path:
            continue

        audited = "No"
        audit_info = fobj.get("audit")
        if audit_info:
            audited = "Si" if audit_info.get("auditado") else "No"
        else:
            marked = fobj.get("audit_marked", "No")
            audited = "Si" if marked in ("Si", "si", "SI") else "No"

        result[abs_path] = {
            "audited": audited,
            "size": fobj.get("size_bytes", fobj.get("Tamano_Bytes", 0)),
            "mtime": fobj.get("modified_at", fobj.get("Fecha_Modificacion", "")),
            "auditor": audit_info.get("auditor", "") if audit_info else "",
        }

    return result


def compare_files(
    current_files: list[FileInfo],
    previous_data: dict[str, dict[str, Any]],
) -> AuditDiff:
    """Compare current scan against previous report data.

    For each file:
    - In previous AND same size + same mtime → keep audit status.
    - In previous BUT different size OR mtime → ``"Modificado"`` (reset to ``"No"``).
    - Not in previous → ``"Nuevo"``.

    Files in previous but not in current are counted as ``removed_files``.
    """
    diff = AuditDiff()
    current_paths = set()

    for f in current_files:
        current_paths.add(f.absolute_path)
        prev = previous_data.get(f.absolute_path)

        if prev is None:
            # New file — not in previous report.
            diff.new_files += 1
            diff.changes[f.absolute_path] = FileChangeStatus(
                status="Nuevo",
                audited="No",
                previous_size=0,
                previous_mtime="",
            )
            continue

        prev_audited = prev.get("audited", "No")
        prev_size = prev.get("size", 0)
        prev_mtime = prev.get("mtime", "")

        # Detect change: both size and mtime must match for "unchanged".
        size_match = prev_size == f.size_bytes
        mtime_match = prev_mtime == f.modified_at

        if size_match and mtime_match:
            # File unchanged.
            if prev_audited == "Si":
                diff.audited_unchanged += 1
                diff.changes[f.absolute_path] = FileChangeStatus(
                    status="Auditado",
                    audited="Si",
                    previous_size=prev_size,
                    previous_mtime=prev_mtime,
                )
            else:
                diff.pending_unchanged += 1
                diff.changes[f.absolute_path] = FileChangeStatus(
                    status="Pendiente",
                    audited="No",
                    previous_size=prev_size,
                    previous_mtime=prev_mtime,
                )
        else:
            # File changed since last report.
            diff.audited_modified += 1
            diff.changes[f.absolute_path] = FileChangeStatus(
                status="Modificado",
                audited="No",
                previous_size=prev_size,
                previous_mtime=prev_mtime,
            )

    # Count removed files (in previous but not in current).
    diff.removed_files = len(set(previous_data.keys()) - current_paths)

    return diff


def calculate_audit_stats(
    current_files: list[FileInfo],
    audit_diff: AuditDiff,
) -> dict[str, Any]:
    """Calculate audit statistics from current files and the diff.

    Returns a dict with per-category file/LOC counts and overall percentages.
    """
    audited_files = 0
    audited_loc = 0
    pending_files = 0
    pending_loc = 0
    modified_files = 0
    modified_loc = 0
    new_files = 0
    new_loc = 0

    for f in current_files:
        status = audit_diff.determine_status(f.absolute_path)
        if status == "Auditado":
            audited_files += 1
            audited_loc += f.code_lines
        elif status == "Pendiente":
            pending_files += 1
            pending_loc += f.code_lines
        elif status == "Modificado":
            modified_files += 1
            modified_loc += f.code_lines
        elif status == "Nuevo":
            new_files += 1
            new_loc += f.code_lines

    total_files = len(current_files)
    total_loc = sum(f.code_lines for f in current_files)
    pct_audited = (audited_loc / total_loc * 100) if total_loc > 0 else 0

    return {
        "total_files": total_files,
        "total_loc": total_loc,
        "audited_files": audited_files,
        "audited_loc": audited_loc,
        "pending_files": pending_files,
        "pending_loc": pending_loc,
        "modified_files": modified_files,
        "modified_loc": modified_loc,
        "new_files": new_files,
        "new_loc": new_loc,
        "removed_files": audit_diff.removed_files,
        "pct_audited": pct_audited,
        "pct_files_audited": (audited_files / total_files * 100) if total_files > 0 else 0,
    }


def apply_audit_to_files(
    files: list[FileInfo],
    audit_diff: AuditDiff,
) -> list[FileInfo]:
    """Return a new list of FileInfo with ``audit_status`` and ``audit_marked`` set.

    Since ``FileInfo`` is frozen, we use ``dataclasses.replace`` to create
    updated copies. Files not in the diff get default ``"Nuevo"`` / ``"No"``.
    """
    from dataclasses import replace

    updated: list[FileInfo] = []
    for f in files:
        status = audit_diff.determine_status(f.absolute_path)
        change = audit_diff.changes.get(f.absolute_path)
        marked = change.audited if change else "No"

        # For "Modificado" files: the auditor mark resets to "No" (needs re-audit).
        if status == "Modificado":
            marked = "No"

        updated.append(replace(f, audit_status=status, audit_marked=marked))
    return updated


# ---------------------------------------------------------------------------
# Backward-compat API (legacy JSON state file — still importable)
# ---------------------------------------------------------------------------

def load_audit_state(filepath: str = ".to-codigo-audit.json") -> AuditState:
    """Load audit state from a JSON file (legacy).

    Returns an empty :class:`AuditState` if the file does not exist.
    """
    state = AuditState()
    if not os.path.isfile(filepath):
        return state
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data: dict[str, Any] = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        logger.warning("Could not load audit state from %s: %s", filepath, e)
        return state

    state.last_updated = data.get("last_updated", "")
    raw_entries = data.get("entries", {})
    for path, entry_data in raw_entries.items():
        state.entries[path] = AuditEntry(
            filepath=path,
            audited=entry_data.get("audited", False),
            auditor=entry_data.get("auditor", ""),
            audit_date=entry_data.get("audit_date", ""),
            notes=entry_data.get("notes", ""),
        )
    return state


def save_audit_state(
    state: AuditState,
    filepath: str = ".to-codigo-audit.json",
) -> None:
    """Save audit state to a JSON file (legacy)."""
    payload: dict[str, Any] = {
        "last_updated": state.last_updated,
        "entries": {
            path: {
                "filepath": entry.filepath,
                "audited": entry.audited,
                "auditor": entry.auditor,
                "audit_date": entry.audit_date,
                "notes": entry.notes,
            }
            for path, entry in sorted(state.entries.items())
        },
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    logger.info("Audit state saved: %s (%d entries)", filepath, len(state.entries))


def mark_file(
    state: AuditState,
    filepath: str,
    audited: bool,
    auditor: str = "",
    notes: str = "",
) -> None:
    """Mark a single file as audited or unaudited (legacy)."""
    abs_path = os.path.abspath(filepath)
    state.mark(abs_path, audited=audited, auditor=auditor, notes=notes)


def merge_audit_state(
    state: AuditState,
    files: list[FileInfo],
) -> AuditState:
    """Ensure every scanned file exists in the audit state (legacy)."""
    for f in files:
        if f.absolute_path not in state.entries:
            state.entries[f.absolute_path] = AuditEntry(
                filepath=f.absolute_path,
                audited=False,
            )
    return state
